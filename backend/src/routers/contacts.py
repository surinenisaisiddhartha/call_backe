import io
import csv
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy.orm import Session
from src.db import get_db, Contact, UploadBatch, CallAttempt, ScheduledCallback, School, CounselorActivity, Appointment, Counselor, CallCostSnapshot
from src.routers.auth import get_current_user
from src.profile import profile_dict, completeness
import openpyxl
import phonenumbers

router = APIRouter(prefix="/api/contacts", tags=["Contacts"])


def resolve_school_id(db: Session, current_user: dict, requested_school_id: str = None) -> str | None:
    """
    Tenant scoping: school users are ALWAYS pinned to their own school —
    whatever they request. Platform admins may target a specific school, and
    default to the original single-tenant school so pre-multitenancy behavior
    is unchanged for them.
    """
    if current_user.get("school_id"):
        return current_user["school_id"]
    if requested_school_id:
        return requested_school_id
    default = db.query(School).filter(School.slug == "shri-ram-academy").first()
    return default.id if default else None


def _normalize_phone(raw_phone) -> str | None:
    """
    Parse/validate a phone number with libphonenumber instead of a
    length-guessing heuristic. Defaults to India ("IN") as the region for
    numbers with no country code, since that was the prior hardcoded
    assumption (+91). Returns E.164 (e.g. "+919876543210") or None if the
    number isn't a real, valid number.
    """
    raw = str(raw_phone).strip()
    try:
        parsed = phonenumbers.parse(raw, "IN")
    except phonenumbers.NumberParseException:
        return None
    if not phonenumbers.is_valid_number(parsed):
        return None
    return phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)


def _extract_row(name_val, phone_val, email_val, notes_val, row_idx: int, dnc_numbers: set):
    """
    Shared validation for one CSV/Excel row (used by both parsing branches
    below). Returns (contact_dict, None) on success or (None, error_dict) on
    a skippable row-level problem.
    """
    if not name_val:
        return None, {"row": row_idx, "error": "Name is empty"}
    if not phone_val:
        return None, {"row": row_idx, "error": "Phone number is empty"}

    clean_phone = _normalize_phone(phone_val)
    if not clean_phone:
        return None, {"row": row_idx, "error": f"'{phone_val}' is not a valid phone number"}

    if clean_phone in dnc_numbers:
        return None, {"row": row_idx, "error": f"Phone number {clean_phone} is on DoNotCall list"}

    return {
        "name": str(name_val).strip(),
        "phone": clean_phone,
        "email": str(email_val).strip() if email_val else None,
        "notes": str(notes_val).strip() if notes_val else None
    }, None


@router.post("/upload")
async def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel (.xlsx/.xls) or CSV file.")

    try:
        contents = await file.read()
        contacts = []
        errors = []

        is_csv = file.filename.endswith('.csv')

        # Fetch the Do-Not-Call list once instead of one query per row
        # (the old code ran a fresh SELECT per row — an N+1 pattern that
        # meant a 5,000-row file issued 5,000 individual DNC lookups).
        dnc_numbers = {row[0] for row in db.query(Contact.phone_number).filter(Contact.status == "DoNotCall").all()}

        if is_csv:
            # --- CSV Parsing ---
            decoded = contents.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(decoded))
            if not reader.fieldnames:
                raise HTTPException(status_code=400, detail="CSV file is empty or has no headers")

            # Map headers
            headers = {h.lower().replace(" ", ""): h for h in reader.fieldnames}
            name_key = next((headers[k] for k in headers if "name" in k), None)
            phone_key = next((headers[k] for k in headers if "phone" in k), None)
            email_key = next((headers[k] for k in headers if "email" in k), None)
            notes_key = next((headers[k] for k in headers if "note" in k), None)

            if not name_key or not phone_key:
                raise HTTPException(
                    status_code=400,
                    detail="Required columns 'Name' and 'PhoneNumber' not found in CSV header row."
                )

            for row_idx, row in enumerate(reader, start=2):
                name_val = row.get(name_key)
                phone_val = row.get(phone_key)
                email_val = row.get(email_key) if email_key else None
                notes_val = row.get(notes_key) if notes_key else None

                contact, error = _extract_row(name_val, phone_val, email_val, notes_val, row_idx, dnc_numbers)
                if error:
                    errors.append(error)
                    continue
                contacts.append(contact)
        else:
            # --- Excel Parsing (existing logic) ---
            workbook = openpyxl.load_workbook(filename=io.BytesIO(contents))
            worksheet = workbook.active
            if worksheet.max_row <= 1:
                raise HTTPException(status_code=400, detail="Excel sheet is empty")

            name_col_idx = None
            phone_col_idx = None
            email_col_idx = None
            notes_col_idx = None

            for col_idx in range(1, worksheet.max_column + 1):
                cell_val = worksheet.cell(row=1, column=col_idx).value
                if cell_val:
                    val = str(cell_val).lower().replace(" ", "")
                    if "name" in val:
                        name_col_idx = col_idx
                    elif "phone" in val:
                        phone_col_idx = col_idx
                    elif "email" in val:
                        email_col_idx = col_idx
                    elif "note" in val:
                        notes_col_idx = col_idx

            if name_col_idx is None or phone_col_idx is None:
                raise HTTPException(
                    status_code=400,
                    detail="Required columns 'Name' and 'PhoneNumber' not found in the header row."
                )

            for row_idx in range(2, worksheet.max_row + 1):
                name_val = worksheet.cell(row=row_idx, column=name_col_idx).value
                phone_val = worksheet.cell(row=row_idx, column=phone_col_idx).value
                email_val = worksheet.cell(row=row_idx, column=email_col_idx).value if email_col_idx else None
                notes_val = worksheet.cell(row=row_idx, column=notes_col_idx).value if notes_col_idx else None

                contact, error = _extract_row(name_val, phone_val, email_val, notes_val, row_idx, dnc_numbers)
                if error:
                    errors.append(error)
                    continue
                contacts.append(contact)

        if not contacts:
            raise HTTPException(status_code=400, detail="No valid rows found in file")

        # Insert upload batch record, pinned to the uploader's school
        school_id = resolve_school_id(db, current_user)
        batch = UploadBatch(
            school_id=school_id,
            file_name=file.filename,
            total_contacts=len(contacts),
            uploaded_by=current_user.get("email", "admin")
        )
        db.add(batch)
        db.commit()
        db.refresh(batch)

        # Insert contacts with automatic Round-Robin assignment
        counselors = db.query(Counselor).filter(Counselor.school_id == school_id).all()
        counselor_count = len(counselors)
        
        success_count = 0
        for idx, c in enumerate(contacts):
            try:
                assigned_id = None
                if counselor_count > 0:
                    assigned_id = counselors[idx % counselor_count].id

                new_contact = Contact(
                    school_id=school_id,
                    batch_id=batch.id,
                    name=c["name"],
                    phone_number=c["phone"],
                    email=c["email"],
                    notes=c["notes"],
                    status="Pending",
                    assigned_counselor_id=assigned_id
                )
                db.add(new_contact)
                success_count += 1
            except Exception as e:
                errors.append({"row": -1, "error": f"Failed to insert {c['name']}: {str(e)}"})
        
        db.commit()

        return {
            "success": True,
            "batchId": batch.id,
            "totalContacts": len(contacts),
            "successCount": success_count,
            "errors": errors
        }

    except HTTPException as he:
        raise he
    except Exception as e:
        print("File parse error:", e)
        raise HTTPException(status_code=500, detail="Failed to process file")


class ManualContactInput(BaseModel):
    name: str
    phone_number: str
    email: str | None = None
    notes: str | None = None

class CreateManualCampaignPayload(BaseModel):
    campaign_name: str
    contacts: list[ManualContactInput]
    start_immediately: bool = False

@router.post("/campaigns/manual")
def create_manual_campaign(
    payload: CreateManualCampaignPayload,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Creates a new outreach campaign manually without needing a CSV file.
    Accepts campaign name, initial contacts, and auto-assigns counselors.
    """
    if not payload.campaign_name.strip():
        raise HTTPException(status_code=400, detail="Campaign name is required")
    if not payload.contacts:
        raise HTTPException(status_code=400, detail="At least one lead is required to create a campaign")

    school_id = resolve_school_id(db, current_user)
    
    batch = UploadBatch(
        school_id=school_id,
        file_name=payload.campaign_name.strip(),
        total_contacts=len(payload.contacts),
        uploaded_by=current_user.get("email", "admin"),
        status="running" if payload.start_immediately else "idle"
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)

    counselors = db.query(Counselor).filter(Counselor.school_id == school_id).all()
    counselor_count = len(counselors)

    created_contacts = []
    for idx, c in enumerate(payload.contacts):
        assigned_id = None
        if counselor_count > 0:
            assigned_id = counselors[idx % counselor_count].id

        new_contact = Contact(
            school_id=school_id,
            batch_id=batch.id,
            name=c.name.strip(),
            phone_number=c.phone_number.strip(),
            email=c.email.strip() if c.email else None,
            notes=c.notes.strip() if c.notes else None,
            status="Pending",
            assigned_counselor_id=assigned_id
        )
        db.add(new_contact)
        created_contacts.append(new_contact)

    db.commit()

    return {
        "success": True,
        "batchId": batch.id,
        "campaignName": batch.file_name,
        "totalContacts": len(created_contacts),
        "message": f"Manual campaign '{batch.file_name}' created with {len(created_contacts)} leads!"
    }

@router.get("/batches")
def get_batches(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from sqlalchemy import func
    query = db.query(UploadBatch)
    if current_user.get("school_id"):
        query = query.filter(UploadBatch.school_id == current_user["school_id"])
    batches = query.order_by(UploadBatch.uploaded_at.desc()).all()

    # One grouped query for EVERY campaign's status counts, instead of one per
    # campaign. Against a remote database the per-campaign version spent nearly
    # all its time on network round trips, and got linearly worse as campaigns
    # were added.
    counts_by_batch: dict = {}
    batch_ids = [b.id for b in batches]
    if batch_ids:
        rows = (
            db.query(Contact.batch_id, Contact.status, func.count(Contact.id))
            .filter(Contact.batch_id.in_(batch_ids))
            .group_by(Contact.batch_id, Contact.status)
            .all()
        )
        for batch_id, status, n in rows:
            counts_by_batch.setdefault(batch_id, {})[status] = n

    result = []
    for b in batches:
        counts = counts_by_batch.get(b.id, {})
        result.append({
            "id": b.id,
            "file_name": b.file_name,
            "uploaded_at": b.uploaded_at.isoformat(),
            "total_contacts": b.total_contacts,
            "uploaded_by": b.uploaded_by,
            "status": b.status or "idle",
            "stats": {
                "pending": counts.get("Pending", 0),
                "calling": counts.get("Calling", 0),
                "completed": counts.get("Completed", 0),
                "needs_reschedule": counts.get("NeedsReschedule", 0),
                "scheduled": counts.get("Scheduled", 0),
                "failed": counts.get("Failed", 0),
            }
        })
    return result


@router.delete("/batches/{batch_id}")
def delete_batch(
    batch_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    batch = db.query(UploadBatch).filter(UploadBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if current_user.get("school_id") and batch.school_id != current_user["school_id"]:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    # Preserve immutable call attempts and billing cost snapshots before deleting campaign contacts
    contacts = db.query(Contact).filter(Contact.batch_id == batch_id).all()
    for c in contacts:
        db.query(CallAttempt).filter(CallAttempt.contact_id == c.id).update({
            "school_id": c.school_id,
            "contact_name": c.name,
            "contact_phone": c.phone_number,
            "contact_id": None
        }, synchronize_session=False)
        db.query(ScheduledCallback).filter(ScheduledCallback.contact_id == c.id).delete(synchronize_session=False)

    db.delete(batch)
    db.commit()
    return {"success": True, "message": "Campaign deleted"}

@router.get("/batches/{batch_id}/stats")
def get_batch_stats(
    batch_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    from sqlalchemy import func
    batch = db.query(UploadBatch).filter(UploadBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if current_user.get("school_id") and batch.school_id != current_user["school_id"]:
        raise HTTPException(status_code=404, detail="Campaign not found")

    status_counts = db.query(Contact.status, func.count(Contact.id)).filter(
        Contact.batch_id == batch_id
    ).group_by(Contact.status).all()
    counts = {s: c for s, c in status_counts}
    total = sum(counts.values()) or 1

    return {
        "id": batch.id,
        "file_name": batch.file_name,
        "status": batch.status or "idle",
        "uploaded_at": batch.uploaded_at.isoformat(),
        "total": sum(counts.values()),
        "pending": counts.get("Pending", 0),
        "calling": counts.get("Calling", 0),
        "completed": counts.get("Completed", 0),
        "needs_reschedule": counts.get("NeedsReschedule", 0),
        "scheduled": counts.get("Scheduled", 0),
        "failed": counts.get("Failed", 0),
        "completion_rate": round(counts.get("Completed", 0) / total * 100, 1),
        "answer_rate": round((counts.get("Completed", 0) + counts.get("Scheduled", 0)) / total * 100, 1),
    }


@router.get("/batches/{batch_id}/history")
def get_batch_history(
    batch_id: str,
    page: int = None,
    page_size: int = 20,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Return call attempts for contacts in this campaign with optional server-side pagination."""
    contact_q = db.query(Contact.id).filter(Contact.batch_id == batch_id)
    if current_user.get("school_id"):
        contact_q = contact_q.filter(Contact.school_id == current_user["school_id"])
    contact_ids = [c.id for c in contact_q.all()]
    if not contact_ids:
        if page is not None:
            return {"items": [], "total": 0, "page": page, "page_size": page_size, "pages": 0}
        return []
    
    query = db.query(CallAttempt, Contact).join(
        Contact, CallAttempt.contact_id == Contact.id
    ).filter(
        CallAttempt.contact_id.in_(contact_ids)
    ).order_by(CallAttempt.started_at.desc())

    total = query.count()
    campaign_name = db.query(UploadBatch.file_name).filter(UploadBatch.id == batch_id).scalar()

    if page is not None:
        p = max(1, page)
        ps = max(1, min(page_size, 100))
        attempts = query.offset((p - 1) * ps).limit(ps).all()
        items = [{
            "id": a.id,
            "contact_id": a.contact_id,
            "contact_name": c.name,
            "contact_phone": c.phone_number,
            "retell_call_id": a.retell_call_id,
            "attempt_number": a.attempt_number,
            "outcome": a.outcome,
            "duration_sec": a.duration_sec,
            "recording_url": a.recording_url,
            "transcript": a.transcript,
            "summary": a.summary,
            "callback_raw_text": a.callback_raw_text,
            "started_at": a.started_at.isoformat() if a.started_at else None,
            "ended_at": a.ended_at.isoformat() if a.ended_at else None,
            "batch_id": batch_id,
            "campaign_name": campaign_name
        } for a, c in attempts]
        return {
            "items": items,
            "total": total,
            "page": p,
            "page_size": ps,
            "pages": (total + ps - 1) // ps
        }

    attempts = query.all()
    return [{
        "id": a.id,
        "contact_id": a.contact_id,
        "contact_name": c.name,
        "contact_phone": c.phone_number,
        "retell_call_id": a.retell_call_id,
        "attempt_number": a.attempt_number,
        "outcome": a.outcome,
        "duration_sec": a.duration_sec,
        "recording_url": a.recording_url,
        "transcript": a.transcript,
        "summary": a.summary,
        "callback_raw_text": a.callback_raw_text,
        "started_at": a.started_at.isoformat() if a.started_at else None,
        "ended_at": a.ended_at.isoformat() if a.ended_at else None,
        "batch_id": batch_id,
        "campaign_name": campaign_name
    } for a, c in attempts]


@router.get("/history/all")
def get_all_call_history(
    status: str = None,
    batch_id: str = None,
    page: int = None,
    page_size: int = 20,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Global call history across all campaigns with server-side pagination."""
    query = db.query(CallAttempt, Contact).join(
        Contact, CallAttempt.contact_id == Contact.id
    )
    if current_user.get("school_id"):
        query = query.filter(Contact.school_id == current_user["school_id"])
    if status:
        query = query.filter(CallAttempt.outcome == status)
    if batch_id:
        query = query.filter(Contact.batch_id == batch_id)

    total = query.count()
    query = query.order_by(CallAttempt.started_at.desc())

    if page is not None:
        p = max(1, page)
        ps = max(1, min(page_size, 100))
        results = query.offset((p - 1) * ps).limit(ps).all()
    else:
        results = query.limit(500).all()

    batch_names = {}
    for a, c in results:
        if c.batch_id and c.batch_id not in batch_names:
            name = db.query(UploadBatch.file_name).filter(UploadBatch.id == c.batch_id).scalar()
            batch_names[c.batch_id] = name or c.batch_id

    import json as _json_ser

    def _parse_analysis_json(raw):
        if not raw:
            return None
        try:
            parsed = _json_ser.loads(raw)
            return parsed if isinstance(parsed, dict) and parsed else None
        except (ValueError, TypeError):
            return None

    items = [{
        "id": a.id,
        "contact_id": a.contact_id,
        "contact_name": c.name,
        "contact_phone": c.phone_number,
        "provider": a.provider or "retell",
        "provider_call_id": a.provider_call_id or a.retell_call_id,
        "retell_call_id": a.retell_call_id or a.provider_call_id,
        "provider_status": a.provider_status,
        "internal_status": a.internal_status,
        "attempt_number": a.attempt_number,
        "outcome": a.outcome,
        "duration_sec": a.duration_sec,
        "recording_url": a.recording_url,
        "transcript": a.transcript,
        "summary": a.summary,
        "callback_raw_text": a.callback_raw_text,
        "user_sentiment": a.user_sentiment,
        "call_successful": a.call_successful,
        "detected_topics": [x for x in (a.detected_topics or "").split(",") if x],
        "analysis": _parse_analysis_json(a.analysis_json),
        "started_at": a.started_at.isoformat() if a.started_at else None,
        "ended_at": a.ended_at.isoformat() if a.ended_at else None,
        "batch_id": c.batch_id,
        "campaign_name": batch_names.get(c.batch_id, c.batch_id) if c.batch_id else None
    } for a, c in results]

    if page is not None:
        return {
            "items": items,
            "total": total,
            "page": p,
            "page_size": ps,
            "pages": (total + ps - 1) // ps
        }

    return items


# ── LLM Enum Normalization ───────────────────────────────────────────────
# Retell's post-call analysis LLM is instructed to output exact enum values
# (e.g. "Hot", "Serious", "Parent"), but LLMs occasionally drift to synonyms
# like "High", "Very Interested", "Mother". Without normalization, these
# unmapped values fall through to default scores (typically 30/40), silently
# under-scoring genuinely interested leads.
#
# Each map covers the canonical values (identity-mapped for clarity) plus
# every observed or plausible synonym. Case-insensitive lookup.

_INTEREST_SYNONYMS: dict[str, str] = {
    # Canonical
    "hot": "Hot", "warm": "Warm", "cold": "Cold", "unclear": "Unclear",
    # Common LLM drift
    "high": "Hot", "very high": "Hot", "very interested": "Hot",
    "strong": "Hot", "keen": "Hot", "eager": "Hot",
    "medium": "Warm", "moderate": "Warm", "somewhat interested": "Warm",
    "interested": "Warm", "mild": "Warm", "curious": "Warm",
    "low": "Cold", "none": "Cold", "not interested": "Cold",
    "no interest": "Cold", "disinterested": "Cold", "negative": "Cold",
    "unknown": "Unclear", "n/a": "Unclear", "na": "Unclear",
    "indeterminate": "Unclear", "cannot determine": "Unclear",
}

_ENGAGEMENT_SYNONYMS: dict[str, str] = {
    # Canonical
    "serious": "Serious", "casual": "Casual",
    "notinterested": "NotInterested", "unclear": "Unclear",
    # Common LLM drift
    "engaged": "Serious", "very engaged": "Serious", "highly engaged": "Serious",
    "active": "Serious", "genuine": "Serious", "substantive": "Serious",
    "light": "Casual", "brief": "Casual", "superficial": "Casual",
    "browsing": "Casual", "polite": "Casual", "pleasant": "Casual",
    "not interested": "NotInterested", "disengaged": "NotInterested",
    "uninterested": "NotInterested", "hostile": "NotInterested",
    "unknown": "Unclear", "n/a": "Unclear", "na": "Unclear",
    "too short": "Unclear", "garbled": "Unclear",
}

_CALLER_TYPE_SYNONYMS: dict[str, str] = {
    # Canonical
    "parent": "Parent", "student": "Student", "other": "Other",
    "notavailable": "NotAvailable", "wrongnumber": "WrongNumber",
    # Common LLM drift
    "mother": "Parent", "father": "Parent", "guardian": "Parent",
    "mom": "Parent", "dad": "Parent", "uncle": "Parent", "aunt": "Parent",
    "grandparent": "Parent", "relative": "Parent",
    "child": "Student", "kid": "Student", "applicant": "Student",
    "agent": "Other", "consultant": "Other", "teacher": "Other",
    "staff": "Other", "friend": "Other", "neighbor": "Other",
    "not available": "NotAvailable", "unavailable": "NotAvailable",
    "busy": "NotAvailable", "away": "NotAvailable",
    "wrong number": "WrongNumber", "wrong person": "WrongNumber",
    "invalid": "WrongNumber",
    "unknown": "Other", "n/a": "Other", "na": "Other",
}


def normalize_analysis_enums(analysis: dict) -> dict:
    """Normalize LLM-produced enum values to canonical vocabulary in-place.

    If the LLM drifts (e.g. outputs "High" instead of "Hot", "Mother"
    instead of "Parent"), this maps the value to its canonical form so
    scoring functions see the expected keys. Unrecognised values pass
    through unchanged — the scoring functions' .get() fallback handles them.
    """
    if not analysis:
        return analysis

    for field, synonyms in (
        ("interest_level",     _INTEREST_SYNONYMS),
        ("engagement_quality", _ENGAGEMENT_SYNONYMS),
        ("caller_type",        _CALLER_TYPE_SYNONYMS),
    ):
        raw = (analysis.get(field) or "").strip()
        if raw:
            canonical = synonyms.get(raw.lower())
            if canonical:
                analysis[field] = canonical

    return analysis


CLASSIFICATIONS = ["HOT", "WARM", "COLD"]

# ── Weighted lead scoring ────────────────────────────────────────────────
# Eight parameters, each scored 0–100 independently, then combined via a
# fixed weight vector. The final weighted score (also 0–100) determines the
# classification:
#
#   75–100   → HOT
#   50–74.99 → WARM
#    0–49.99 → COLD
#
# The percentage determines the classification — never a single field like
# engagement_quality or interest_level on its own.

PARAM_WEIGHTS = {
    "appointment_conversion_intent": 0.25,
    "engagement_quality":            0.20,
    "interest_level":                0.15,
    "conversation_depth":            0.10,
    "caller_relevance":              0.10,
    "admission_intent":              0.10,
    "requirement_fit":               0.05,
    "follow_up_intent":              0.05,
}


def _score_appointment(booked: bool, callback: bool, analysis: dict, contact) -> int:
    """Appointment Conversion Intent — 15%."""
    if not contact:
        return 0
    if contact.status == "DoNotCall":
        return 0
    if booked:
        return 100
    # Check for explicit campus visit / appointment request in analysis
    if analysis:
        next_step = (analysis.get("recommended_next_step") or "").lower()
        synopsis = (analysis.get("call_synopsis") or "").lower()
        combined = next_step + " " + synopsis
        if any(kw in combined for kw in ("visit", "appointment", "campus tour", "schedule a visit")):
            return 90
    if callback:
        return 75
    # Check for admission info requests in analysis
    if analysis:
        topics = (analysis.get("topics_discussed") or "").lower()
        if any(kw in topics for kw in ("admissions", "fees", "curriculum")):
            return 65
        interest = (analysis.get("interest_level") or "").strip()
        if interest in ("Hot", "Warm"):
            return 50
        engagement = (analysis.get("engagement_quality") or "").strip()
        if engagement in ("Serious", "Casual"):
            return 35
        caller_type = (analysis.get("caller_type") or "").strip()
        if caller_type == "WrongNumber":
            return 0
        return 15
    return None  # Missing signal if no booking, no callback, and no call analysis yet


def _score_engagement(analysis: dict, sentiment: str | None = None) -> int:
    """Engagement Quality — 20%.

    Uses the LLM's engagement_quality as the base score, then adjusts by
    Retell's own user_sentiment when available. A "Negative" sentiment
    caller who the LLM rated as "Serious" was probably serious about their
    OBJECTIONS, not about enrolling — so the score is penalised. A
    "Positive" sentiment gives a small boost.
    """
    if not analysis and not sentiment:
        return None  # missing signal

    if analysis:
        engagement = (analysis.get("engagement_quality") or "").strip()
        base = {"Serious": 100, "Casual": 50, "Unclear": 30, "NotInterested": 0}.get(engagement, 30)
    else:
        # No LLM analysis, but sentiment alone can still provide a signal
        base = 50  # neutral baseline

    # Apply sentiment modifier
    if sentiment:
        s = sentiment.strip().lower()
        if s == "negative":
            base = max(0, base - 20)
        elif s == "positive":
            base = min(100, base + 10)
        # "neutral" → no change

    return base


def _score_interest(analysis: dict) -> int:
    """Interest Level — 15%."""
    if not analysis:
        return None  # missing signal
    interest = (analysis.get("interest_level") or "").strip()
    return {"Hot": 100, "Warm": 65, "Cold": 20, "Unclear": 30}.get(interest, 30)


def _score_conversation_depth(analysis: dict, had_engaged_call: bool, detected_topics: str) -> int:
    """Conversation Depth — 10%.  Pure question-density scoring.

    Scores by the number of UNIQUE topics/questions the caller raised across
    ALL their engaged calls.  Duration is used upstream only as a minimum
    threshold (≥ 10 s) to decide if a call counts as "engaged" — it no longer
    directly influences the score, because a concise parent who asks 3 sharp
    questions in 40 seconds is just as valuable as one who takes 5 minutes.

    Returns None (missing signal) when there is no evidence of a real
    conversation at all.
    """
    # Unique topic count from both LLM analysis and keyword detection
    all_topics: set[str] = set()
    if analysis:
        for t in (analysis.get("topics_discussed") or "").split(","):
            cleaned = t.strip()
            if cleaned and cleaned.lower() != "none":
                all_topics.add(cleaned.lower())
    if detected_topics:
        for t in detected_topics.split(","):
            cleaned = t.strip()
            if cleaned:
                all_topics.add(cleaned.lower())

    topic_count = len(all_topics)

    # No topics AND no engaged call → missing signal
    if topic_count == 0 and not had_engaged_call:
        return None

    if topic_count >= 5:
        return 100
    elif topic_count == 4:
        return 90
    elif topic_count == 3:
        return 80
    elif topic_count == 2:
        return 65
    elif topic_count == 1:
        return 45
    else:
        # Had an engaged call but no topics detected
        return 20


def _score_caller_relevance(analysis: dict) -> int:
    """Caller Relevance — 10%."""
    if not analysis:
        return None  # missing signal
    caller_type = (analysis.get("caller_type") or "").strip()
    return {"Parent": 100, "Student": 75, "Other": 40,
            "NotAvailable": 20, "WrongNumber": 0}.get(caller_type, 40)


def _score_admission_intent(contact) -> int:
    """Admission Intent — 10%. Uses the 20-point profile captured mid-call."""
    score = 0
    signals = 0

    # decision_timeline
    timeline = getattr(contact, "decision_timeline", None)
    if timeline:
        signals += 1
        score += {"Immediate": 100, "ThisMonth": 85, "ThisQuarter": 70,
                  "NextYear": 45, "Unknown": 30}.get(timeline, 30)

    # admission_urgency
    urgency = getattr(contact, "admission_urgency", None)
    if urgency:
        signals += 1
        score += {"Urgent": 100, "Planned": 70, "JustExploring": 45,
                  "Unknown": 30}.get(urgency, 30)

    # campus_visit_interest — positive signal
    visit = getattr(contact, "campus_visit_interest", None)
    if visit:
        signals += 1
        score += {"Yes": 90, "AlreadyVisited": 80, "No": 20,
                  "Unknown": 30}.get(visit, 30)

    # decision_maker — Self/Both means they can act
    dm = getattr(contact, "decision_maker", None)
    if dm:
        signals += 1
        score += {"Self": 90, "Both": 80, "Spouse": 50,
                  "ExtendedFamily": 40, "Unknown": 30}.get(dm, 30)

    # competition_considered — actively comparing = serious
    comp = getattr(contact, "competition_considered", None)
    if comp and comp.lower() not in ("none", "n/a", "na", ""):
        signals += 1
        score += 75  # actively comparing schools is a positive signal

    if signals == 0:
        return None  # no profile data yet — missing signal
    return min(100, score // signals)


def _score_requirement_fit(contact) -> int:
    """Requirement / Fit — 5%. How much useful family info was established."""
    fit_fields = [
        "grade_sought", "academic_year", "board_preference", "locality",
        "current_school", "transport_needed", "boarding_needed",
        "special_requirements", "budget_band",
    ]
    filled = sum(1 for f in fit_fields
                 if getattr(contact, f, None) and str(getattr(contact, f)).lower() not in
                 ("unknown", "n/a", "na", "none"))

    if filled >= 7:
        return 100
    elif filled >= 5:
        return 75
    elif filled >= 3:
        return 50
    elif filled >= 1:
        return 25
    return None  # no info at all — missing signal


def _score_follow_up(callback: bool, contact, analysis: dict) -> int:
    """Follow-up Intent — 5%."""
    score = 0
    signals = 0

    if callback:
        signals += 1
        score += 100  # clear next action

    pref_time = getattr(contact, "preferred_contact_time", None)
    if pref_time and pref_time.lower() not in ("unknown", "n/a", "na", "none"):
        signals += 1
        score += 80  # willing to set a time

    pref_lang = getattr(contact, "language_preference", None)
    if pref_lang and pref_lang not in ("Unknown",):
        signals += 1
        score += 60  # shared language preference

    if analysis:
        next_step = (analysis.get("recommended_next_step") or "").strip()
        if next_step and next_step.lower() not in ("none", "n/a", ""):
            signals += 1
            score += 80

    if signals == 0:
        return None  # missing signal
    return min(100, score // signals)


def _classify(score: float) -> str:
    """Final classification from the weighted score — the ONLY rule."""
    if score >= 75:
        return "HOT"
    elif score >= 50:
        return "WARM"
    return "COLD"


def _generate_reason(param_scores: dict, weighted_breakdown: dict,
                     classification: str, missing_params: list) -> str:
    """Human-readable explanation of why this lead got its classification."""
    parts = []

    # Highlight the strongest contributors
    sorted_params = sorted(weighted_breakdown.items(), key=lambda x: x[1], reverse=True)
    top = [p for p, w in sorted_params[:3] if w > 0]

    label_map = {
        "appointment_conversion_intent": "conversion intent",
        "engagement_quality": "engagement",
        "interest_level": "interest",
        "conversation_depth": "conversation depth",
        "caller_relevance": "caller relevance",
        "admission_intent": "admission intent",
        "requirement_fit": "requirement fit",
        "follow_up_intent": "follow-up intent",
    }

    if top:
        strengths = [label_map.get(p, p) for p in top]
        parts.append("Strong in: " + ", ".join(strengths))

    # Note weak areas
    weak = [label_map.get(p, p) for p, s in param_scores.items() if s is not None and s < 30]
    if weak:
        parts.append("Weak in: " + ", ".join(weak[:2]))

    if missing_params:
        parts.append(f"{len(missing_params)} parameter(s) had no data")

    if not parts:
        if classification == "COLD":
            parts.append("Limited engagement and conversion signals")
        elif classification == "WARM":
            parts.append("Moderate engagement and interest signals")
        else:
            parts.append("Strong engagement, interest, and conversion signals")

    return ". ".join(parts) + "."


def compute_interest_levels(db: Session, contacts: list) -> dict:
    """Label-only view; compute_lead_scores has the score and the reasoning."""
    return {cid: v["classification"] for cid, v in compute_lead_scores(db, contacts).items()}


def compute_lead_scores(db: Session, contacts: list) -> dict:
    """
    {contact_id: {lead_score, classification, parameter_scores,
                   weighted_score_breakdown, classification_reason}}

    Eight parameters, each scored 0–100 independently, combined via fixed
    weights. Missing analysis fields are handled by redistributing their
    weight across available parameters so the final score stays 0–100.

    Batched queries for the whole page — never one query per contact.
    """
    import json as _json
    from src.db import Appointment

    ids = [c.id for c in contacts]
    if not ids:
        return {}

    booked = {
        r[0] for r in db.query(Appointment.contact_id).filter(
            Appointment.contact_id.in_(ids), Appointment.status == "Booked"
        ).all()
    }
    # ONLY callbacks the caller asked for. Auto-retries (call_type "Reminder")
    # don't count — they reward not picking up the phone.
    callbacks = {
        r[0] for r in db.query(ScheduledCallback.contact_id).filter(
            ScheduledCallback.contact_id.in_(ids),
            ScheduledCallback.status.in_(["Scheduled", "Triggering"]),
            ScheduledCallback.call_type != "Reminder",
        ).all()
    }

    # ── Cumulative best-intent aggregation across ALL calls ─────────────
    # Instead of keeping only the last call's analysis (which let a brief
    # follow-up call erase rich signals from an earlier conversation), we
    # now merge the BEST intent signal for each field across all engaged
    # calls, and UNION all topics.  Duration is only a minimum-threshold
    # filter (≥ 10 s = "engaged"), not a scoring factor.
    #
    # Rankings for "best" per enum field (index 0 = highest intent):
    _INTEREST_RANK = {"Hot": 0, "Warm": 1, "Cold": 2, "Unclear": 3}
    _ENGAGEMENT_RANK = {"Serious": 0, "Casual": 1, "Unclear": 2, "NotInterested": 3}
    _CALLER_RANK = {"Parent": 0, "Student": 1, "Other": 2, "NotAvailable": 3, "WrongNumber": 4}

    merged_analysis: dict[str, dict] = {}   # contact_id → best-of analysis dict
    merged_topics: dict[str, set] = {}      # contact_id → union of all detected topics
    had_engaged: dict[str, bool] = {}       # contact_id → True if ≥ 1 engaged call
    merged_sentiment: dict[str, str] = {}   # contact_id → best (most positive) sentiment

    _NON_ENGAGED = {"NoAnswer", "Busy", "Failed", "IncompleteHangup"}
    rows = (
        db.query(
            CallAttempt.contact_id, CallAttempt.analysis_json,
            CallAttempt.duration_sec, CallAttempt.detected_topics,
            CallAttempt.outcome, CallAttempt.user_sentiment,
        )
        .filter(CallAttempt.contact_id.in_(ids))
        .order_by(CallAttempt.started_at.asc())
        .all()
    )

    _SENTIMENT_RANK = {"Positive": 0, "Neutral": 1, "Negative": 2}

    def _pick_best_enum(current: str | None, new: str | None, ranking: dict) -> str | None:
        """Return whichever value ranks higher (lower index = higher intent)."""
        if not new:
            return current
        if not current:
            return new
        return current if ranking.get(current, 99) <= ranking.get(new, 99) else new

    def _pick_longest(current: str | None, new: str | None) -> str | None:
        """Keep the longer string (carries more detail)."""
        if not new:
            return current
        if not current:
            return new
        return current if len(current) >= len(new) else new

    for contact_id, raw, duration, det_topics, outcome, sentiment_val in rows:
        engaged = outcome not in _NON_ENGAGED and (duration or 0) >= 10

        # Merge sentiment: keep the most positive across all calls
        if sentiment_val:
            merged_sentiment[contact_id] = _pick_best_enum(
                merged_sentiment.get(contact_id), sentiment_val, _SENTIMENT_RANK
            )
        if engaged:
            had_engaged[contact_id] = True

        # Accumulate detected topics (keyword-based) across ALL engaged calls
        if engaged and det_topics:
            if contact_id not in merged_topics:
                merged_topics[contact_id] = set()
            for t in det_topics.split(","):
                if t.strip():
                    merged_topics[contact_id].add(t.strip())

        # Merge LLM analysis: keep the highest-intent value per field
        if raw:
            try:
                parsed = _json.loads(raw)
                if not isinstance(parsed, dict) or not parsed:
                    continue
            except (ValueError, TypeError):
                continue

            # Normalize LLM enum drift BEFORE merging (e.g. "High" → "Hot")
            normalize_analysis_enums(parsed)

            if contact_id not in merged_analysis:
                merged_analysis[contact_id] = {}
            best = merged_analysis[contact_id]

            # Enum fields: keep the highest-intent value
            best["interest_level"] = _pick_best_enum(
                best.get("interest_level"), parsed.get("interest_level"), _INTEREST_RANK)
            best["engagement_quality"] = _pick_best_enum(
                best.get("engagement_quality"), parsed.get("engagement_quality"), _ENGAGEMENT_RANK)
            best["caller_type"] = _pick_best_enum(
                best.get("caller_type"), parsed.get("caller_type"), _CALLER_RANK)

            # String fields: keep the longest (most detailed) value
            for str_field in ("call_synopsis", "recommended_next_step", "concerns_raised"):
                best[str_field] = _pick_longest(best.get(str_field), parsed.get(str_field))

            # Topics: union across all calls
            new_topics = (parsed.get("topics_discussed") or "").split(",")
            existing = set(
                t.strip() for t in (best.get("topics_discussed") or "").split(",")
                if t.strip() and t.strip().lower() != "none"
            )
            for t in new_topics:
                cleaned = t.strip()
                if cleaned and cleaned.lower() != "none":
                    existing.add(cleaned)
            best["topics_discussed"] = ", ".join(sorted(existing)) if existing else ""

            # Primary topic: keep from the call with highest engagement
            if parsed.get("primary_topic") and not best.get("primary_topic"):
                best["primary_topic"] = parsed["primary_topic"]

    attempted_contact_ids = {r.contact_id for r in rows}
    completed_contact_ids = {
        r.contact_id for r in rows
        if r.outcome is not None or r.duration_sec is not None or r.analysis_json is not None
    }

    out = {}
    for c in contacts:
        has_completed_interaction = (c.id in completed_contact_ids) or (c.id in booked) or (c.id in callbacks)

        # If call is live right now (Calling) and no completed call has arrived yet:
        if (c.status == "Calling" or (c.id in attempted_contact_ids and c.id not in completed_contact_ids)) and not (c.id in booked or c.id in callbacks):
            out[c.id] = {
                "score": None,
                "classification": "CALLING",
                "reasons": ["Call is actively in progress — scoring activates upon call completion."],
                "parameter_scores": {k: 0 for k in PARAM_WEIGHTS},
                "weighted_score_breakdown": {k: 0.0 for k in PARAM_WEIGHTS},
                "classification_reason": "Call is actively in progress — scoring activates upon call completion.",
                "missing_params": list(PARAM_WEIGHTS.keys()),
                "uncontacted": True,
            }
            continue

        # Leads with NO completed call attempts, NO appointments, and NO callbacks are UNCONTACTED
        if not has_completed_interaction:
            out[c.id] = {
                "score": None,
                "classification": "UNSCORED",
                "reasons": ["Uncontacted lead — scoring activates after the first call attempt."],
                "parameter_scores": {k: 0 for k in PARAM_WEIGHTS},
                "weighted_score_breakdown": {k: 0.0 for k in PARAM_WEIGHTS},
                "classification_reason": "Uncontacted lead — scoring activates after the first call attempt.",
                "qualification_source": [],
                "qualification_reason": "Uncontacted lead",
                "missing_params": list(PARAM_WEIGHTS.keys()),
                "uncontacted": True,
            }
            continue

        analysis = merged_analysis.get(c.id)
        det_topics = ",".join(merged_topics.get(c.id, set()))
        engaged = had_engaged.get(c.id, False)
        is_booked = c.id in booked
        is_callback = c.id in callbacks

        # ── DNC / Permanent Suppression Hard Rule Override ───────────────
        # An explicit DNC request or wrong number overrides any numerical score.
        is_dnc = (
            getattr(c, "status", None) in ("DoNotCall", "DNC", "Unqualified") or
            (analysis and analysis.get("caller_type") == "WrongNumber") or
            (analysis and "do not call" in (analysis.get("call_synopsis") or "").lower())
        )
        if is_dnc:
            out[c.id] = {
                "score": 0.0,
                "classification": "DO_NOT_CALL",
                "reasons": ["Explicit do-not-call request / contact suppression override."],
                "parameter_scores": {k: 0 for k in PARAM_WEIGHTS},
                "weighted_score_breakdown": {k: 0.0 for k in PARAM_WEIGHTS},
                "classification_reason": "Explicit do-not-call request / contact suppression override",
                "qualification_source": ["DNC Override"],
                "qualification_reason": "Caller requested Do Not Call or number is invalid",
                "missing_params": [],
            }
            continue

        # Compute each parameter (None = missing signal)
        raw_scores = {
            "appointment_conversion_intent": _score_appointment(is_booked, is_callback, analysis, c),
            "engagement_quality":            _score_engagement(analysis, merged_sentiment.get(c.id)),
            "interest_level":                _score_interest(analysis),
            "conversation_depth":            _score_conversation_depth(analysis, engaged, det_topics),
            "caller_relevance":              _score_caller_relevance(analysis),
            "admission_intent":              _score_admission_intent(c),
            "requirement_fit":               _score_requirement_fit(c),
            "follow_up_intent":              _score_follow_up(is_callback, c, analysis),
        }

        # Separate available vs missing
        available = {k: v for k, v in raw_scores.items() if v is not None}
        missing = [k for k, v in raw_scores.items() if v is None]

        # Normalize weights so available parameters sum to 1.0
        if available:
            total_available_weight = sum(PARAM_WEIGHTS[k] for k in available)
            if total_available_weight > 0:
                scale = 1.0 / total_available_weight
            else:
                scale = 1.0

            weighted_breakdown = {}
            for k, v in available.items():
                weighted_breakdown[k] = round(v * PARAM_WEIGHTS[k] * scale, 2)
            # Fill missing with 0 for display
            for k in missing:
                weighted_breakdown[k] = 0.0

            final_score = round(sum(weighted_breakdown.values()), 2)
        else:
            # No data at all — score is 0
            weighted_breakdown = {k: 0.0 for k in PARAM_WEIGHTS}
            final_score = 0.0

        final_score = max(0.0, min(100.0, final_score))
        classification = _classify(final_score)

        # Parameter scores for display (None → 0 in output)
        param_scores = {k: (v if v is not None else 0) for k, v in raw_scores.items()}

        reason = _generate_reason(param_scores, weighted_breakdown, classification, missing)

        # Derive qualification sources (why lead was prioritized)
        sources = []
        if is_booked:
            sources.append("Campus visit / appointment booked")
        if is_callback:
            sources.append("Counselor callback requested")
        if getattr(c, "grade_sought", None):
            sources.append(f"Target grade: {c.grade_sought}")
        if getattr(c, "budget_range", None):
            sources.append(f"Budget: {c.budget_range}")
        if merged_sentiment.get(c.id) == "Positive":
            sources.append("High positive sentiment")
        if not sources:
            sources.append("General admission inquiry")

        qual_reason = reason
        if classification == "COLD":
            qual_reason = "Low immediate intent / future inquiry timeline"

        out[c.id] = {
            "score": final_score,
            "classification": classification,
            "reasons": [reason],
            "parameter_scores": param_scores,
            "weighted_score_breakdown": weighted_breakdown,
            "classification_reason": reason,
            "qualification_source": sources,
            "qualification_reason": qual_reason,
            "missing_params": missing,
        }
    return out


def rescore_contact(db: Session, contact_id: str) -> dict | None:
    """Computes and writes fresh lead_score and lead_classification to the Contact table."""
    contact = db.query(Contact).filter(Contact.id == contact_id).first()
    if not contact:
        return None
    scores = compute_lead_scores(db, [contact])
    if contact_id in scores:
        data = scores[contact_id]
        contact.lead_score = data.get("score")
        contact.lead_classification = data.get("classification")
        db.commit()
        db.refresh(contact)
        return data
    return None


def persist_lead_scores(db: Session, contacts: list) -> dict:
    """Batched helper to compute and persist lead scores on multiple contacts."""
    if not contacts:
        return {}
    scores = compute_lead_scores(db, contacts)
    for c in contacts:
        if c.id in scores:
            c.lead_score = scores[c.id].get("score")
            c.lead_classification = scores[c.id].get("classification")
    db.commit()
    return scores


def auto_assign_hot_lead_to_available_counselor(db: Session, contact: Contact) -> Counselor | None:
    """
    Auto-assigns an unassigned qualified lead (HOT, WARM, or Callback requested)
    to the available Counselor who currently has the fewest active leads and is not
    at max capacity. Returns the assigned Counselor or None.
    """
    if contact.assigned_counselor_id:
        return None

    classification = (contact.lead_classification or "COLD").upper()
    is_qualified = (
        classification in ("HOT", "WARM") or 
        (contact.lead_score or 0) >= 40 or 
        contact.status in ("NeedsReschedule", "Scheduled") or
        contact.counselor_followup_status == "Pending"
    )
    if not is_qualified:
        return None

    from sqlalchemy import func
    counselors = db.query(Counselor).filter(
        Counselor.school_id == contact.school_id,
        Counselor.availability_status == "Available"
    ).all()
    if not counselors:
        return None

    # Build active-lead counts (counselor follow-up not completed)
    lead_counts = dict(
        db.query(Contact.assigned_counselor_id, func.count(Contact.id))
        .filter(
            Contact.school_id == contact.school_id,
            Contact.assigned_counselor_id.isnot(None),
            Contact.counselor_followup_status != "Completed"
        )
        .group_by(Contact.assigned_counselor_id)
        .all()
    )

    # Filter out at-capacity counselors, then pick the one with fewest leads
    eligible = [
        c for c in counselors
        if lead_counts.get(c.id, 0) < (c.max_capacity or 50)
    ]
    if not eligible:
        return None

    best = min(eligible, key=lambda c: lead_counts.get(c.id, 0))
    contact.assigned_counselor_id = best.id
    contact.counselor_followup_status = "Pending"

    # Log as activity
    activity = CounselorActivity(
        contact_id=contact.id,
        counselor_id=best.id,
        action_type="AutoAssign",
        outcome=f"{classification} Lead Auto-Assigned",
        notes=f"Auto-assigned to {best.name} (active leads: {lead_counts.get(best.id, 0)})"
    )
    db.add(activity)

    try:
        from src.events import event_manager
        event_manager.broadcast_sync(
            "COUNSELOR_ASSIGNED",
            {"contact_id": contact.id, "counselor_id": best.id, "counselor_name": best.name},
            school_id=contact.school_id
        )
    except Exception:
        pass

    return best

auto_assign_hot_lead = auto_assign_hot_lead_to_available_counselor


def persist_lead_scores(db: Session, contacts: list) -> dict:
    """
    Recompute and STORE the score for these contacts.

    Called whenever something that feeds a score changes — a call is analysed,
    an appointment is booked, a callback is requested. Storing it is what makes
    "show me the hottest leads" an indexed SQL query over one page instead of a
    Python pass over every lead the school has ever had.

    Deliberately scoped to the contacts handed in. Never recompute the whole
    table: at 10,000 leads a day that is the exact operation that would make
    the dashboard unusable.
    """
    from datetime import datetime as _dt

    scored = compute_lead_scores(db, contacts)
    now = _dt.utcnow()
    for c in contacts:
        s = scored.get(c.id)
        if not s:
            continue
        if s.get("uncontacted"):
            if c.lead_score is not None or c.lead_classification != "UNSCORED":
                c.lead_score = None
                c.lead_classification = "UNSCORED"
                c.lead_scored_at = None
        else:
            if c.lead_score != s["score"] or c.lead_classification != s["classification"]:
                c.lead_score = s["score"]
                c.lead_classification = s["classification"]
                c.lead_scored_at = now
            # Automatically assign HOT/WARM leads to counselors if unassigned
            if (s.get("classification") in ("HOT", "WARM") or (s.get("score") or 0) >= 50) and not c.assigned_counselor_id:
                auto_assign_hot_lead_to_available_counselor(db, c)
    db.commit()
    return scored


def rescore_contact(db: Session, contact_id: str):
    """Rescore one contact after a call, booking or callback. Never raises —
    a scoring failure must not break the webhook that triggered it."""
    try:
        contact = db.query(Contact).filter(Contact.id == contact_id).first()
        if contact:
            persist_lead_scores(db, [contact])
    except Exception as e:
        print(f"[SCORING] Could not rescore contact {contact_id}: {e}")


class CreateContactPayload(BaseModel):
    name: str
    phone_number: str
    email: str | None = None
    notes: str | None = None
    batch_id: str | None = None
    assigned_counselor_id: str | None = None
    
    # Optional profile fields
    child_name: str | None = None
    child_age: str | None = None
    grade_sought: str | None = None
    academic_year: str | None = None
    board_preference: str | None = None
    locality: str | None = None
    current_school: str | None = None
    transport_needed: str | None = None
    budget_band: str | None = None
    admission_urgency: str | None = None
    decision_timeline: str | None = None


@router.post("")
def create_contact(
    payload: CreateContactPayload,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Manually create a new contact/lead with validation, normalization,
    and optional initial profile information.
    """
    if not payload.name or not payload.name.strip():
        raise HTTPException(status_code=400, detail="Contact name is required.")
    
    if not payload.phone_number or not payload.phone_number.strip():
        raise HTTPException(status_code=400, detail="Phone number is required.")
    
    clean_phone = _normalize_phone(payload.phone_number)
    if not clean_phone:
        raise HTTPException(status_code=400, detail=f"'{payload.phone_number}' is not a valid phone number. Please enter a valid 10-digit number.")

    school_id = resolve_school_id(db, current_user)

    # Check Do-Not-Call list
    dnc_exists = db.query(Contact).filter(
        Contact.phone_number == clean_phone,
        Contact.status == "DoNotCall",
        Contact.school_id == school_id if school_id else True
    ).first()
    if dnc_exists:
        raise HTTPException(status_code=400, detail=f"Phone number {clean_phone} is on the Do Not Call list.")

    # Check duplicate in same school
    existing = db.query(Contact).filter(
        Contact.phone_number == clean_phone,
        Contact.school_id == school_id if school_id else True
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"A contact with phone number {clean_phone} already exists ({existing.name}).")

    assigned_cns_id = payload.assigned_counselor_id
    if assigned_cns_id and assigned_cns_id.lower() in ("none", "", "null"):
        assigned_cns_id = None

    # If no counselor explicitly selected, try round-robin across available counselors
    if not assigned_cns_id and school_id:
        counselors = db.query(Counselor).filter(Counselor.school_id == school_id, Counselor.availability_status == "Available").all()
        if counselors:
            from sqlalchemy import func
            lead_counts = dict(
                db.query(Contact.assigned_counselor_id, func.count(Contact.id))
                .filter(
                    Contact.school_id == school_id,
                    Contact.assigned_counselor_id.isnot(None),
                    Contact.status != "Completed"
                )
                .group_by(Contact.assigned_counselor_id)
                .all()
            )
            eligible = [c for c in counselors if lead_counts.get(c.id, 0) < (c.max_capacity or 50)]
            if eligible:
                best = min(eligible, key=lambda c: lead_counts.get(c.id, 0))
                assigned_cns_id = best.id

    new_contact = Contact(
        school_id=school_id,
        batch_id=payload.batch_id if payload.batch_id else None,
        name=payload.name.strip(),
        phone_number=clean_phone,
        email=payload.email.strip() if payload.email else None,
        notes=payload.notes.strip() if payload.notes else None,
        status="Pending",
        assigned_counselor_id=assigned_cns_id,
        # Profile fields
        child_name=payload.child_name.strip() if payload.child_name else None,
        child_age=payload.child_age.strip() if payload.child_age else None,
        grade_sought=payload.grade_sought.strip() if payload.grade_sought else None,
        academic_year=payload.academic_year.strip() if payload.academic_year else None,
        board_preference=payload.board_preference.strip() if payload.board_preference else None,
        locality=payload.locality.strip() if payload.locality else None,
        current_school=payload.current_school.strip() if payload.current_school else None,
        transport_needed=payload.transport_needed.strip() if payload.transport_needed else None,
        budget_band=payload.budget_band.strip() if payload.budget_band else None,
        admission_urgency=payload.admission_urgency.strip() if payload.admission_urgency else None,
        decision_timeline=payload.decision_timeline.strip() if payload.decision_timeline else None,
    )
    db.add(new_contact)
    db.commit()
    db.refresh(new_contact)

    # Initial lead score
    persist_lead_scores(db, [new_contact])
    db.refresh(new_contact)

    if assigned_cns_id:
        activity = CounselorActivity(
            contact_id=new_contact.id,
            counselor_id=assigned_cns_id,
            action_type="AutoAssign" if not payload.assigned_counselor_id else "ManualAssign",
            outcome="Lead Created",
            notes="Manually entered lead"
        )
        db.add(activity)
        db.commit()

    return {
        "success": True,
        "contact": {
            "id": new_contact.id,
            "name": new_contact.name,
            "phone_number": new_contact.phone_number,
            "email": new_contact.email,
            "status": new_contact.status,
            "lead_score": new_contact.lead_score,
            "lead_classification": new_contact.lead_classification,
            "assigned_counselor_id": new_contact.assigned_counselor_id,
        }
    }


@router.get("")
def get_contacts(
    status: str = None,
    batchId: str = None,
    search: str = None,
    interest: str = None,
    counselor_id: str = None,
    counselor_followup_status: str = None,
    admission_urgency: str = None,
    page: int = 1,
    page_size: int = 50,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    A PAGE of leads, ranked best-first.

    Filtering, ranking and paging all happen in SQL against the stored
    lead_score.
    """
    query = db.query(Contact)
    if current_user.get("school_id"):
        query = query.filter(Contact.school_id == current_user["school_id"])
    if status:
        if status == "active":
            query = query.filter(Contact.status != "Completed")
        else:
            query = query.filter(Contact.status == status)
    if batchId:
        query = query.filter(Contact.batch_id == batchId)
    if counselor_followup_status:
        norm_followup = counselor_followup_status.strip()
        if norm_followup.lower() == "active":
            query = query.filter(Contact.counselor_followup_status != "Completed")
        elif norm_followup.lower() == "completed":
            query = query.filter(Contact.counselor_followup_status == "Completed")
        elif norm_followup.lower() in ("pending", "inprogress"):
            query = query.filter(Contact.counselor_followup_status == norm_followup)
        else:
            query = query.filter(Contact.counselor_followup_status == norm_followup)
    if interest:
        norm_interest = interest.strip().upper()
        if norm_interest == "HOT":
            query = query.filter((Contact.lead_classification == "HOT") | (Contact.lead_score >= 75))
        elif norm_interest == "WARM":
            query = query.filter((Contact.lead_classification == "WARM") | ((Contact.lead_score >= 50) & (Contact.lead_score < 75)))
        elif norm_interest == "COLD":
            query = query.filter(
                (Contact.lead_classification == "COLD") | ((Contact.lead_score < 50) & (Contact.lead_score.isnot(None)))
            ).filter(
                Contact.lead_classification != "HOT",
                Contact.lead_classification != "WARM",
                (Contact.lead_score < 50) | (Contact.lead_score.is_(None))
            )
        elif norm_interest == "UNSCORED":
            query = query.filter((Contact.lead_classification == "UNSCORED") | (Contact.lead_score.is_(None)))
        else:
            query = query.filter(Contact.lead_classification == interest.strip())
    if counselor_id:
        if counselor_id == "unassigned":
            query = query.filter(Contact.assigned_counselor_id == None)
        else:
            query = query.filter(Contact.assigned_counselor_id == counselor_id)
    if admission_urgency:
        query = query.filter(Contact.admission_urgency == admission_urgency)
    if search:
        query = query.filter(
            Contact.name.ilike(f"%{search}%") |
            Contact.phone_number.ilike(f"%{search}%")
        )

    total = query.count()

    page = max(1, page)
    page_size = max(1, min(page_size, 200))   # a caller cannot ask for everything
    rows = (
        query.order_by(Contact.lead_score.desc(), Contact.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    # Only this page is scored, so the reasons stay fresh without touching the
    # rest of the table.
    scored = compute_lead_scores(db, rows)

    # Synchronise any score discrepancies back to DB
    needs_commit = False
    for c in rows:
        s = scored.get(c.id)
        if s:
            target_class = s.get("classification")
            target_score = s.get("score")
            if target_class and (c.lead_classification != target_class or c.lead_score != target_score):
                c.lead_classification = target_class
                c.lead_score = target_score
                needs_commit = True
    if needs_commit:
        try:
            db.commit()
        except Exception:
            pass

    # Batch compute per-lead call costs and duration
    contact_ids = [c.id for c in rows]
    cost_by_contact = {}
    if contact_ids:
        from sqlalchemy import func
        cost_rows = db.query(
            CallAttempt.contact_id,
            func.count(CallAttempt.id),
            func.coalesce(func.sum(CallCostSnapshot.provider_total_cost), 0.0),
            func.coalesce(func.sum(CallCostSnapshot.customer_billable_total), 0.0),
            func.coalesce(func.sum(CallAttempt.duration_sec), 0.0)
        ).outerjoin(
            CallCostSnapshot, CallAttempt.id == CallCostSnapshot.call_attempt_id
        ).filter(
            CallAttempt.contact_id.in_(contact_ids)
        ).group_by(CallAttempt.contact_id).all()

        for cid, call_cnt, prov_cost, cust_billed, tot_dur in cost_rows:
            mins = tot_dur / 60.0 if tot_dur else 0.0
            # Fallback estimation if snapshots not yet generated for legacy calls
            if prov_cost == 0.0 and tot_dur > 0:
                prov_cost = round(mins * 6.68, 2)
            if cust_billed == 0.0 and tot_dur > 0:
                cust_billed = round(mins * 15.00, 2)

            margin = max(0.0, float(cust_billed) - float(prov_cost))
            margin_pct = round((margin / float(cust_billed) * 100.0), 1) if cust_billed > 0 else 0.0

            cost_by_contact[cid] = {
                "call_count": int(call_cnt),
                "provider_total_cost": round(float(prov_cost), 2),
                "customer_billable_total": round(float(cust_billed), 2),
                "net_margin": round(margin, 2),
                "gross_margin_percent": margin_pct,
                "total_duration_sec": round(float(tot_dur), 1),
            }

    is_admin = current_user.get("role") == "admin" or not current_user.get("school_id")

    items = [
        {
            "id": c.id,
            "school_id": c.school_id,
            "batch_id": c.batch_id,
            "name": c.name,
            "phone_number": c.phone_number,
            "email": c.email,
            "notes": c.notes,
            "status": c.status,
            "interest_level": scored.get(c.id, {}).get("classification", c.lead_classification or "UNSCORED"),
            "lead_classification": scored.get(c.id, {}).get("classification", c.lead_classification or "UNSCORED"),
            "lead_score": scored.get(c.id, {}).get("score", c.lead_score),
            "score_reasons": scored.get(c.id, {}).get("reasons", []),
            "parameter_scores": scored.get(c.id, {}).get("parameter_scores", {}),
            "weighted_score_breakdown": scored.get(c.id, {}).get("weighted_score_breakdown", {}),
            "classification_reason": scored.get(c.id, {}).get("classification_reason", ""),
            "call_cost": {
                "call_count": cost_by_contact.get(c.id, {}).get("call_count", 0),
                "customer_billable_total": cost_by_contact.get(c.id, {}).get("customer_billable_total", 0.0),
                "total_duration_sec": cost_by_contact.get(c.id, {}).get("total_duration_sec", 0.0),
                "provider_total_cost": cost_by_contact.get(c.id, {}).get("provider_total_cost") if is_admin else None,
                "net_margin": cost_by_contact.get(c.id, {}).get("net_margin") if is_admin else None,
                "gross_margin_percent": cost_by_contact.get(c.id, {}).get("gross_margin_percent") if is_admin else None,
            },
            # Only the points actually learned, plus the count — a counselor
            # scanning the queue wants "14/20 known", not twenty nulls.
            "profile": profile_dict(c),
            "profile_completeness": completeness(c),
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None,
            "assigned_counselor_id": c.assigned_counselor_id,
            "counselor_followup_status": c.counselor_followup_status or "Pending",
        }
        for c in rows
    ]

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": (total + page_size - 1) // page_size,
    }


@router.get("/stats")
def get_contact_stats(
    batchId: str = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Counts by status for the whole tenant — the numbers on the dashboard tiles
    and the campaign summary.

    These used to be derived on the client by fetching every contact and
    calling .filter() over the array. Once GET /contacts became paginated that
    became both wrong and unfixable from the client: the page cap is 200, so a
    school with 10,000 leads would have shown "200 Total Leads" and a
    "Completed" count drawn from whichever 200 happened to be on page one.
    Wrong numbers that look plausible are worse than a crash, so the counts are
    computed in SQL where the whole table is visible.

    MUST STAY DEFINED ABOVE GET /{id} — FastAPI matches in declaration order,
    and /{id} would otherwise swallow "stats" as a contact id.
    """
    from sqlalchemy import func

    query = db.query(Contact.status, func.count(Contact.id))
    if current_user.get("school_id"):
        query = query.filter(Contact.school_id == current_user["school_id"])
    if batchId:
        query = query.filter(Contact.batch_id == batchId)

    by_status = {status or "Unknown": count for status, count in query.group_by(Contact.status).all()}

    class_query = db.query(Contact.lead_classification, func.count(Contact.id))
    if current_user.get("school_id"):
        class_query = class_query.filter(Contact.school_id == current_user["school_id"])
    if batchId:
        class_query = class_query.filter(Contact.batch_id == batchId)
    by_class = {c or "UNSCORED": count for c, count in class_query.group_by(Contact.lead_classification).all()}

    followup_query = db.query(Contact.counselor_followup_status, func.count(Contact.id))
    if current_user.get("school_id"):
        followup_query = followup_query.filter(Contact.school_id == current_user["school_id"])
    if batchId:
        followup_query = followup_query.filter(Contact.batch_id == batchId)
    by_followup = {f or "Pending": count for f, count in followup_query.group_by(Contact.counselor_followup_status).all()}

    return {
        "total": sum(by_status.values()),
        "by_status": by_status,
        "by_classification": by_class,
        "by_followup_status": by_followup,
        # Spelled out so the client never has to know the status vocabulary.
        "completed": by_status.get("Completed", 0),
        "calling": by_status.get("Calling", 0),
        "pending": by_status.get("Pending", 0),
        "needs_reschedule": by_status.get("NeedsReschedule", 0),
        "scheduled": by_status.get("Scheduled", 0),
        "failed": by_status.get("Failed", 0),
        "hot": by_class.get("HOT", 0),
        "warm": by_class.get("WARM", 0),
        "cold": by_class.get("COLD", 0),
        "unscored": by_class.get("UNSCORED", 0),
        "followup_pending": by_followup.get("Pending", 0) + by_followup.get("InProgress", 0),
        "followup_completed": by_followup.get("Completed", 0),
    }


from src.db import Counselor

@router.get("/counselors/all")
def get_counselors(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    from sqlalchemy import func
    school_id = resolve_school_id(db, current_user)
    counselors = db.query(Counselor).filter(Counselor.school_id == school_id).all()

    # Compute active lead counts per counselor in one query
    lead_counts = dict(
        db.query(Contact.assigned_counselor_id, func.count(Contact.id))
        .filter(
            Contact.school_id == school_id,
            Contact.assigned_counselor_id.isnot(None),
            Contact.status != "Completed"
        )
        .group_by(Contact.assigned_counselor_id)
        .all()
    )

    return [
        {
            "id": c.id,
            "name": c.name,
            "email": c.email,
            "phone_number": c.phone_number,
            "availability_status": c.availability_status or "Available",
            "max_capacity": c.max_capacity or 50,
            "active_lead_count": lead_counts.get(c.id, 0),
        }
        for c in counselors
    ]

class CounselorCreatePayload(BaseModel):
    name: str
    email: str
    phone_number: str | None = None

@router.post("/counselors")
def create_counselor(
    payload: CounselorCreatePayload,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    school_id = resolve_school_id(db, current_user)
    
    existing = db.query(Counselor).filter(Counselor.school_id == school_id, Counselor.email == payload.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Counselor with this email already exists")

    temp_password = None
    from src import cognito
    if cognito.cognito_enabled():
        try:
            temp_password = cognito.create_school_user(payload.email, school_id)
        except Exception as e:
            print(f"[COUNSELOR ONBOARD] Cognito account creation failed: {e}")
            raise HTTPException(status_code=400, detail=f"Failed to create login account: {str(e)}")

    new_counselor = Counselor(
        school_id=school_id,
        name=payload.name,
        email=payload.email,
        phone_number=payload.phone_number
    )
    db.add(new_counselor)
    db.commit()
    db.refresh(new_counselor)
    return {
        "success": True, 
        "counselor": {"id": new_counselor.id, "name": new_counselor.name, "email": new_counselor.email},
        "temp_password": temp_password
    }

@router.delete("/counselors/{id}")
def delete_counselor(
    id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    school_id = resolve_school_id(db, current_user)
    counselor = db.query(Counselor).filter(Counselor.id == id, Counselor.school_id == school_id).first()
    if not counselor:
        raise HTTPException(status_code=404, detail="Counselor not found")
    
    from src import cognito
    if cognito.cognito_enabled():
        try:
            cognito.delete_school_user(counselor.email)
        except Exception as e:
            print(f"[COUNSELOR DELETION] Cognito account deletion failed: {e}")

    db.delete(counselor)
    db.commit()
    return {"success": True, "message": "Counselor removed successfully"}

@router.post("/counselors/auto-assign")
def auto_assign_contacts(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    from src.events import event_manager
    school_id = resolve_school_id(db, current_user)
    
    counselors = db.query(Counselor).filter(Counselor.school_id == school_id).all()
    if not counselors:
        raise HTTPException(status_code=400, detail="No onboarded counselors available for assignment")
    
    # Priority: filter for available counselors first
    available_counselors = [c for c in counselors if c.availability_status == "Available"]
    if not available_counselors:
        available_counselors = counselors

    active_counselor_ids = {c.id for c in counselors}

    # Pick up (1) unassigned AND (2) assigned to deleted counselor
    all_contacts = db.query(Contact).filter(Contact.school_id == school_id).all()
    needs_assignment = [
        c for c in all_contacts
        if c.assigned_counselor_id is None
        or c.assigned_counselor_id not in active_counselor_ids
    ]
    
    if not needs_assignment:
        return {"success": True, "message": "All contacts are already assigned to active counselors", "assigned_count": 0}
        
    # Sort priority: HOT leads first, then WARM, then others
    def _lead_priority(c: Contact):
        if c.lead_classification == "HOT" or (c.lead_score or 0) >= 75:
            return 0
        if c.lead_classification == "WARM" or (c.lead_score or 0) >= 50:
            return 1
        return 2

    needs_assignment.sort(key=_lead_priority)

    cns_count = len(available_counselors)
    for idx, contact in enumerate(needs_assignment):
        chosen = available_counselors[idx % cns_count]
        contact.assigned_counselor_id = chosen.id
        contact.counselor_followup_status = contact.counselor_followup_status or "Pending"
        
        # Log activity
        activity = CounselorActivity(
            contact_id=contact.id,
            counselor_id=chosen.id,
            action_type="Assignment",
            outcome="Auto-Assigned",
            notes=f"Auto-assigned to {chosen.name} ({chosen.availability_status})"
        )
        db.add(activity)

    db.commit()

    # Broadcast real-time SSE updates for all affected leads
    for contact in needs_assignment:
        event_manager.broadcast_sync(
            "COUNSELOR_ASSIGNED",
            {"contact_id": contact.id, "counselor_id": contact.assigned_counselor_id},
            school_id=school_id
        )

    return {
        "success": True,
        "message": f"Successfully auto-assigned {len(needs_assignment)} leads across {cns_count} active counselors",
        "assigned_count": len(needs_assignment)
    }


class BulkUpdatePayload(BaseModel):
    contact_ids: list[str]
    action: str  # 'assign', 'auto_assign', 'status', 'status_change', 'schedule'
    assigned_counselor_id: str | None = None
    status: str | None = None
    counselor_followup_status: str | None = None
    scheduled_for: str | None = None


@router.post("/bulk-update")
def bulk_update_contacts(
    payload: BulkUpdatePayload,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    from src.events import event_manager
    school_id = resolve_school_id(db, current_user)
    
    if not payload.contact_ids:
        raise HTTPException(status_code=400, detail="No contact IDs provided")
        
    contacts = db.query(Contact).filter(
        Contact.id.in_(payload.contact_ids),
        Contact.school_id == school_id
    ).all()
    
    if not contacts:
        raise HTTPException(status_code=404, detail="No matching contacts found")
        
    counselors_map = {c.id: c for c in db.query(Counselor).filter(Counselor.school_id == school_id).all()}
    
    if payload.action == "assign":
        target_cns = counselors_map.get(payload.assigned_counselor_id) if payload.assigned_counselor_id else None
        for c in contacts:
            c.assigned_counselor_id = target_cns.id if target_cns else None
            if target_cns:
                c.counselor_followup_status = c.counselor_followup_status or "Pending"
                activity = CounselorActivity(
                    contact_id=c.id,
                    counselor_id=target_cns.id,
                    action_type="Assignment",
                    outcome="Assigned to Counselor",
                    notes=f"Assigned to counselor {target_cns.name}"
                )
                db.add(activity)
            else:
                c.counselor_followup_status = "Pending"
        db.commit()
        for c in contacts:
            event_manager.broadcast_sync("COUNSELOR_ASSIGNED", {"contact_id": c.id, "counselor_id": c.assigned_counselor_id}, school_id=school_id)
        return {"success": True, "message": f"Successfully assigned {len(contacts)} contacts"}

    elif payload.action == "auto_assign":
        available_cns = [c for c in counselors_map.values() if c.availability_status == "Available"] or list(counselors_map.values())
        if not available_cns:
            raise HTTPException(status_code=400, detail="No active counselors available for auto-assignment")
            
        cns_count = len(available_cns)
        for idx, c in enumerate(contacts):
            chosen = available_cns[idx % cns_count]
            c.assigned_counselor_id = chosen.id
            c.counselor_followup_status = c.counselor_followup_status or "Pending"
            activity = CounselorActivity(
                contact_id=c.id,
                counselor_id=chosen.id,
                action_type="Assignment",
                outcome="Auto-Assigned",
                notes=f"Auto-assigned to {chosen.name}"
            )
            db.add(activity)
        db.commit()
        for c in contacts:
            event_manager.broadcast_sync("COUNSELOR_ASSIGNED", {"contact_id": c.id, "counselor_id": c.assigned_counselor_id}, school_id=school_id)
        return {"success": True, "message": f"Auto-assigned {len(contacts)} contacts across {cns_count} counselors"}

    elif payload.action in ("status", "status_change"):
        st = payload.status or "Completed"
        for c in contacts:
            if st in ("Pending", "Calling", "Completed", "NeedsReschedule", "Scheduled", "Failed"):
                c.status = st
            if st in ("Completed", "InProgress", "Pending") or payload.counselor_followup_status:
                c.counselor_followup_status = payload.counselor_followup_status or st
            activity = CounselorActivity(
                contact_id=c.id,
                counselor_id=c.assigned_counselor_id,
                action_type="StatusChange",
                outcome=f"Status set to {st}",
                notes="Bulk status updated by counselor / admin"
            )
            db.add(activity)
        db.commit()
        for c in contacts:
            event_manager.broadcast_sync("CONTACT_UPDATED", {"contact_id": c.id, "status": c.status}, school_id=school_id)
        return {"success": True, "message": f"Updated status for {len(contacts)} contacts"}

    elif payload.action == "schedule":
        if not payload.scheduled_for:
            raise HTTPException(status_code=400, detail="scheduled_for datetime is required")
        try:
            dt = datetime.fromisoformat(payload.scheduled_for.replace("Z", "+00:00")).replace(tzinfo=None)
        except Exception:
            dt = datetime.utcnow()

        for c in contacts:
            c.status = "Scheduled"
            sched = ScheduledCallback(
                contact_id=c.id,
                scheduled_for=dt,
                status="Scheduled"
            )
            db.add(sched)
            activity = CounselorActivity(
                contact_id=c.id,
                counselor_id=c.assigned_counselor_id,
                action_type="Call",
                outcome="Callback Scheduled",
                notes=f"Bulk callback scheduled for {dt.strftime('%d %b %Y, %I:%M %p')}"
            )
            db.add(activity)
        db.commit()
        for c in contacts:
            event_manager.broadcast_sync("CALLBACK_SCHEDULED", {"contact_id": c.id, "scheduled_for": dt.isoformat()}, school_id=school_id)
        return {"success": True, "message": f"Scheduled callbacks for {len(contacts)} contacts"}

    else:
        raise HTTPException(status_code=400, detail=f"Unsupported bulk action: {payload.action}")


# ── Counselor Availability Toggle ──────────────────────────────────────

class CounselorUpdatePayload(BaseModel):
    availability_status: str | None = None  # Available, InConsultation, OnLeave
    max_capacity: int | None = None
    name: str | None = None
    phone_number: str | None = None

@router.patch("/counselors/{counselor_id}")
def update_counselor(
    counselor_id: str,
    payload: CounselorUpdatePayload,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    school_id = resolve_school_id(db, current_user)
    counselor = db.query(Counselor).filter(Counselor.id == counselor_id, Counselor.school_id == school_id).first()
    if not counselor:
        raise HTTPException(status_code=404, detail="Counselor not found")
    if payload.availability_status is not None:
        if payload.availability_status not in ("Available", "InConsultation", "OnLeave"):
            raise HTTPException(status_code=400, detail="availability_status must be Available, InConsultation, or OnLeave")
        counselor.availability_status = payload.availability_status
    if payload.max_capacity is not None:
        counselor.max_capacity = max(1, payload.max_capacity)
    if payload.name is not None:
        counselor.name = payload.name
    if payload.phone_number is not None:
        counselor.phone_number = payload.phone_number
    db.commit()
    return {"success": True, "counselor": {"id": counselor.id, "availability_status": counselor.availability_status, "max_capacity": counselor.max_capacity}}


# ── Counselor Performance Analytics ────────────────────────────────────

@router.get("/counselors/analytics")
def get_counselor_analytics(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    from sqlalchemy import func
    school_id = resolve_school_id(db, current_user)
    counselors = db.query(Counselor).filter(Counselor.school_id == school_id).all()
    if not counselors:
        return []

    results = []
    for c in counselors:
        total_assigned = db.query(func.count(Contact.id)).filter(
            Contact.assigned_counselor_id == c.id
        ).scalar() or 0

        completed_count = db.query(func.count(Contact.id)).filter(
            Contact.assigned_counselor_id == c.id,
            Contact.counselor_followup_status == "Completed"
        ).scalar() or 0

        hot_leads = db.query(func.count(Contact.id)).filter(
            Contact.assigned_counselor_id == c.id,
            Contact.lead_classification == "HOT"
        ).scalar() or 0

        active_leads = db.query(func.count(Contact.id)).filter(
            Contact.assigned_counselor_id == c.id,
            Contact.counselor_followup_status != "Completed"
        ).scalar() or 0

        # Activity count in last 7 days
        from datetime import timedelta
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        activity_count_7d = db.query(func.count(CounselorActivity.id)).filter(
            CounselorActivity.counselor_id == c.id,
            CounselorActivity.created_at >= seven_days_ago
        ).scalar() or 0

        # Average response time: first activity per contact by this counselor
        # vs contact.created_at
        avg_response_hours = None
        try:
            from sqlalchemy import text
            row = db.execute(text("""
                SELECT AVG(EXTRACT(EPOCH FROM (first_activity - contact_created)) / 3600.0)
                FROM (
                    SELECT ca.contact_id,
                           MIN(ca.created_at) AS first_activity,
                           c.created_at AS contact_created
                    FROM counselor_activities ca
                    JOIN contacts c ON c.id = ca.contact_id
                    WHERE ca.counselor_id = :cid
                    GROUP BY ca.contact_id, c.created_at
                ) sub
            """), {"cid": c.id}).fetchone()
            if row and row[0] is not None:
                avg_response_hours = round(float(row[0]), 1)
        except Exception:
            pass

        conversion_rate = round((completed_count / total_assigned * 100), 1) if total_assigned > 0 else 0.0

        results.append({
            "counselor_id": c.id,
            "name": c.name,
            "email": c.email,
            "availability_status": c.availability_status or "Available",
            "total_assigned": total_assigned,
            "completed_count": completed_count,
            "conversion_rate": conversion_rate,
            "hot_leads_assigned": hot_leads,
            "active_leads": active_leads,
            "activity_count_7d": activity_count_7d,
            "avg_response_hours": avg_response_hours,
        })

    # Sort by conversion rate desc
    results.sort(key=lambda x: x["conversion_rate"], reverse=True)
    return results





# ── Structured Follow-Up Activities ────────────────────────────────────

class ActivityPayload(BaseModel):
    action_type: str        # Call, WhatsApp, Email, CampusVisit, Note
    outcome: str | None = None
    notes: str | None = None

@router.post("/{contact_id}/activities")
def log_activity(
    contact_id: str,
    payload: ActivityPayload,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    contact = db.query(Contact).filter(Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    if current_user.get("school_id") and contact.school_id != current_user["school_id"]:
        raise HTTPException(status_code=404, detail="Contact not found")

    # Resolve the logged-in user's counselor id (if any)
    counselor_id = None
    email = current_user.get("email")
    if email:
        counselor = db.query(Counselor).filter(
            Counselor.email == email,
            Counselor.school_id == contact.school_id
        ).first()
        if counselor:
            counselor_id = counselor.id

    activity = CounselorActivity(
        contact_id=contact_id,
        counselor_id=counselor_id,
        action_type=payload.action_type,
        outcome=payload.outcome,
        notes=payload.notes,
    )
    db.add(activity)
    db.commit()
    db.refresh(activity)
    return {
        "success": True,
        "activity": {
            "id": activity.id,
            "contact_id": activity.contact_id,
            "counselor_id": activity.counselor_id,
            "action_type": activity.action_type,
            "outcome": activity.outcome,
            "notes": activity.notes,
            "created_at": activity.created_at.isoformat() if activity.created_at else None,
        }
    }

@router.get("/{contact_id}/activities")
def get_activities(
    contact_id: str,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    contact = db.query(Contact).filter(Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    if current_user.get("school_id") and contact.school_id != current_user["school_id"]:
        raise HTTPException(status_code=404, detail="Contact not found")

    from sqlalchemy import func
    total = db.query(func.count(CounselorActivity.id)).filter(
        CounselorActivity.contact_id == contact_id
    ).scalar() or 0

    rows = db.query(CounselorActivity).filter(
        CounselorActivity.contact_id == contact_id
    ).order_by(CounselorActivity.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    # Resolve counselor names
    counselor_ids = {a.counselor_id for a in rows if a.counselor_id}
    name_map = {}
    if counselor_ids:
        for c in db.query(Counselor).filter(Counselor.id.in_(counselor_ids)).all():
            name_map[c.id] = c.name

    return {
        "items": [
            {
                "id": a.id,
                "contact_id": a.contact_id,
                "counselor_id": a.counselor_id,
                "counselor_name": name_map.get(a.counselor_id, "System"),
                "action_type": a.action_type,
                "outcome": a.outcome,
                "notes": a.notes,
                "created_at": a.created_at.isoformat() if a.created_at else None,
            }
            for a in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": (total + page_size - 1) // page_size,
    }


# ── Bulk Operations ────────────────────────────────────────────────────

class BulkUpdatePayload(BaseModel):
    contact_ids: list[str]
    action: str              # assign, status_change, schedule
    assigned_counselor_id: str | None = None
    status: str | None = None
    scheduled_for: str | None = None  # ISO datetime

@router.post("/bulk-update")
def bulk_update_contacts(
    payload: BulkUpdatePayload,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    school_id = current_user.get("school_id")
    q = db.query(Contact).filter(Contact.id.in_(payload.contact_ids))
    if school_id:
        q = q.filter(Contact.school_id == school_id)
    contacts = q.all()

    if not contacts:
        raise HTTPException(status_code=404, detail="No matching contacts found")

    updated = 0
    if payload.action == "assign" and payload.assigned_counselor_id:
        target_id = None if payload.assigned_counselor_id.lower() in ("none", "", "null") else payload.assigned_counselor_id
        for c in contacts:
            c.assigned_counselor_id = target_id
            c.updated_at = datetime.utcnow()
            updated += 1
        db.commit()

    elif payload.action == "status_change" and payload.status:
        # Resolve counselor for activity logging
        counselor_id = None
        email = current_user.get("email")
        if email and school_id:
            cns = db.query(Counselor).filter(Counselor.email == email, Counselor.school_id == school_id).first()
            if cns:
                counselor_id = cns.id

        for c in contacts:
            old_status = c.status
            c.status = payload.status
            c.updated_at = datetime.utcnow()
            updated += 1
            # Log activity
            activity = CounselorActivity(
                contact_id=c.id,
                counselor_id=counselor_id,
                action_type="StatusChange",
                outcome=payload.status,
                notes=f"Bulk status change: {old_status} → {payload.status}"
            )
            db.add(activity)
        db.commit()

    elif payload.action == "schedule" and payload.scheduled_for:
        from dateutil.parser import isoparse
        try:
            schedule_dt = isoparse(payload.scheduled_for)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid scheduled_for datetime")

        for c in contacts:
            # Avoid duplicates
            existing = db.query(ScheduledCallback).filter(
                ScheduledCallback.contact_id == c.id,
                ScheduledCallback.status == "Scheduled"
            ).first()
            if not existing:
                cb = ScheduledCallback(
                    contact_id=c.id,
                    scheduled_for=schedule_dt,
                    reason="Bulk schedule",
                    status="Scheduled",
                    call_type="Follow-up"
                )
                db.add(cb)
                updated += 1
        db.commit()
    else:
        raise HTTPException(status_code=400, detail="Invalid action or missing required fields")

    return {"success": True, "updated": updated, "message": f"{updated} contacts updated"}



@router.get("/{id}")
def get_contact_history(
    id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    contact = db.query(Contact).filter(Contact.id == id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    if current_user.get("school_id") and contact.school_id != current_user["school_id"]:
        raise HTTPException(status_code=404, detail="Contact not found")

    attempts = db.query(CallAttempt).filter(CallAttempt.contact_id == id).order_by(CallAttempt.started_at.desc()).all()
    schedules = db.query(ScheduledCallback).filter(ScheduledCallback.contact_id == id).order_by(ScheduledCallback.scheduled_for.desc()).all()
    appointments = db.query(Appointment).filter(Appointment.contact_id == id).order_by(Appointment.scheduled_for.desc()).all()

    # Pre-fetch call cost snapshots for itemized per-call financial auditing
    attempt_ids = [a.id for a in attempts]
    snapshots_by_attempt = {}
    if attempt_ids:
        snaps = db.query(CallCostSnapshot).filter(CallCostSnapshot.call_attempt_id.in_(attempt_ids)).all()
        for s in snaps:
            snapshots_by_attempt[s.call_attempt_id] = s

    # Serialise attempts explicitly so the post-call analysis comes back as a
    # real object rather than the JSON string it is stored as. A malformed or
    # legacy value must not break the whole history view, so parsing failures
    # degrade to no analysis rather than raising.
    import json as _json

    def _attempt_dict(a):
        analysis = None
        if a.analysis_json:
            try:
                parsed = _json.loads(a.analysis_json)
                if isinstance(parsed, dict) and parsed:
                    analysis = parsed
            except (ValueError, TypeError):
                analysis = None

        snap = snapshots_by_attempt.get(a.id)
        dur = float(a.duration_sec or 0.0)
        mins = dur / 60.0
        if snap:
            prov_cost = float(snap.provider_total_cost or 0.0)
            cust_billed = float(snap.customer_billable_total or 0.0)
            markup = float(snap.markup_amount or 0.0)
            margin_pct = float(snap.gross_margin_percent or 0.0)
            source = snap.cost_source or "rate_card"
            curr = snap.currency or "INR"
        else:
            prov_cost = round(mins * 6.68, 2)
            cust_billed = round(mins * 15.00, 2)
            markup = max(0.0, cust_billed - prov_cost)
            margin_pct = round((markup / cust_billed * 100.0), 1) if cust_billed > 0 else 0.0
            source = "estimated"
            curr = "INR"

        is_admin_user = current_user.get("role") == "admin" or not current_user.get("school_id")
        cost_obj = {
            "customer_billable_total": round(cust_billed, 2),
            "currency": curr,
            "provider_total_cost": round(prov_cost, 2) if is_admin_user else None,
            "markup_amount": round(markup, 2) if is_admin_user else None,
            "gross_margin_percent": margin_pct if is_admin_user else None,
            "cost_source": source if is_admin_user else None,
        }

        return {
            "id": a.id,
            "provider": a.provider or "retell",
            "provider_call_id": a.provider_call_id or a.retell_call_id,
            "retell_call_id": a.retell_call_id or a.provider_call_id,
            "provider_status": a.provider_status,
            "internal_status": a.internal_status,
            "attempt_number": a.attempt_number,
            "started_at": a.started_at.isoformat() if a.started_at else None,
            "ended_at": a.ended_at.isoformat() if a.ended_at else None,
            "outcome": a.outcome,
            "duration_sec": a.duration_sec,
            "transcript": a.transcript,
            "summary": a.summary,
            "recording_url": a.recording_url,
            "callback_raw_text": a.callback_raw_text,
            "user_sentiment": a.user_sentiment,
            "call_successful": a.call_successful,
            "analysis": analysis,
            "cost": cost_obj,
            "detected_topics": [x for x in (a.detected_topics or "").split(",") if x],
        }

    def _appointment_dict(apt):
        return {
            "id": apt.id,
            "scheduled_for": apt.scheduled_for.isoformat() if apt.scheduled_for else None,
            "status": apt.status,
            "meeting_type": apt.meeting_type,
            "purpose": apt.purpose,
            "virtual_meeting_link": apt.virtual_meeting_link,
            "created_at": apt.created_at.isoformat() if apt.created_at else None,
        }

    # The lead's standing judgement, so the drawer answers "is this person
    # worth my time?" without the reader having to reconstruct it from a list
    # of calls.
    scored = compute_lead_scores(db, [contact]).get(contact.id, {})

    # Every topic this person has ever raised, across all their calls — one
    # call's topics say what they asked that day, this says what they care about.
    all_topics = []
    for a in attempts:
        for label in (a.detected_topics or "").split(","):
            if label and label not in all_topics:
                all_topics.append(label)

    def _schedule_dict(sch):
        target_label = "Counselor Callback" if sch.call_type == "Counselor" else "AI Agent Callback"
        return {
            "id": sch.id,
            "scheduled_for": sch.scheduled_for.isoformat() if sch.scheduled_for else None,
            "status": sch.status,
            "call_type": sch.call_type or "Follow-up",
            "callback_target": target_label,
            "reason": sch.reason,
            "created_at": sch.created_at.isoformat() if sch.created_at else None,
        }

    # Counselor Activities & Timeline
    activities = db.query(CounselorActivity).filter(CounselorActivity.contact_id == id).order_by(CounselorActivity.created_at.desc()).all()
    counselor_ids = {a.counselor_id for a in activities if a.counselor_id}
    if contact.assigned_counselor_id:
        counselor_ids.add(contact.assigned_counselor_id)
    counselor_map = {}
    if counselor_ids:
        for c in db.query(Counselor).filter(Counselor.id.in_(counselor_ids)).all():
            counselor_map[c.id] = {"id": c.id, "name": c.name, "email": c.email, "phone_number": c.phone_number}

    assigned_counselor = counselor_map.get(contact.assigned_counselor_id) if contact.assigned_counselor_id else None

    def _activity_dict(act):
        return {
            "id": act.id,
            "counselor_id": act.counselor_id,
            "counselor_name": counselor_map.get(act.counselor_id, {}).get("name", "System"),
            "action_type": act.action_type,
            "outcome": act.outcome,
            "notes": act.notes,
            "created_at": act.created_at.isoformat() if act.created_at else None,
        }

    return {
        "contact": contact,
        "lead_score": scored.get("score", 0),
        "classification": scored.get("classification", "COLD"),
        "score_reasons": scored.get("reasons", []),
        "parameter_scores": scored.get("parameter_scores", {}),
        "weighted_score_breakdown": scored.get("weighted_score_breakdown", {}),
        "classification_reason": scored.get("classification_reason", ""),
        "assigned_counselor": assigned_counselor,
        "counselor_followup_status": contact.counselor_followup_status or "Pending",
        # What the agent learned about the family, mid-call. Only the points
        # actually captured — see profile.py.
        "profile": profile_dict(contact),
        "profile_completeness": completeness(contact),
        "topics_asked": all_topics,
        "attempts": [_attempt_dict(a) for a in attempts],
        "schedules": [_schedule_dict(sch) for sch in schedules],
        "appointments": [_appointment_dict(apt) for apt in appointments],
        "activities": [_activity_dict(act) for act in activities],
    }

@router.delete("/{id}")
def delete_contact(
    id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    contact = db.query(Contact).filter(Contact.id == id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    if current_user.get("school_id") and contact.school_id != current_user["school_id"]:
        raise HTTPException(status_code=404, detail="Contact not found")
    
    # Preserve immutable call attempts and billing cost snapshots before deleting contact
    db.query(CallAttempt).filter(CallAttempt.contact_id == id).update({
        "school_id": contact.school_id,
        "contact_name": contact.name,
        "contact_phone": contact.phone_number,
        "contact_id": None
    }, synchronize_session=False)

    # Clean up dependent scheduling records
    db.query(ScheduledCallback).filter(ScheduledCallback.contact_id == id).delete(synchronize_session=False)
    
    db.delete(contact)
    db.commit()
    return {"success": True, "message": "Contact deleted"}



class ContactUpdatePayload(BaseModel):
    notes: str | None = None
    email: str | None = None
    name: str | None = None
    status: str | None = None
    assigned_counselor_id: str | None = None
    counselor_followup_status: str | None = None

@router.patch("/{id}")
def update_contact(
    id: str,
    payload: ContactUpdatePayload,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    contact = db.query(Contact).filter(Contact.id == id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    if current_user.get("school_id") and contact.school_id != current_user["school_id"]:
        raise HTTPException(status_code=404, detail="Contact not found")
    
    if payload.notes is not None:
        contact.notes = payload.notes
    if payload.email is not None:
        contact.email = payload.email
    if payload.name is not None:
        contact.name = payload.name

    # Track status changes and assignment changes as structured activities
    counselor_id = None
    email_user = current_user.get("email")
    if email_user and contact.school_id:
        cns = db.query(Counselor).filter(Counselor.email == email_user, Counselor.school_id == contact.school_id).first()
        if cns:
            counselor_id = cns.id

    if payload.counselor_followup_status is not None:
        old_followup = contact.counselor_followup_status
        contact.counselor_followup_status = payload.counselor_followup_status
        if old_followup != payload.counselor_followup_status:
            activity = CounselorActivity(
                contact_id=contact.id,
                counselor_id=counselor_id,
                action_type="FollowupStatus",
                outcome=payload.counselor_followup_status,
                notes=f"Follow-up status updated: {old_followup or 'Pending'} → {payload.counselor_followup_status}"
            )
            db.add(activity)

    if payload.status is not None:
        old_status = contact.status
        contact.status = payload.status
        if old_status != payload.status:
            activity = CounselorActivity(
                contact_id=contact.id,
                counselor_id=counselor_id,
                action_type="StatusChange",
                outcome=payload.status,
                notes=f"{old_status} → {payload.status}"
            )
            db.add(activity)

    if payload.assigned_counselor_id is not None:
        old_counselor = contact.assigned_counselor_id
        if payload.assigned_counselor_id.lower() in ("none", "", "null"):
            contact.assigned_counselor_id = None
        else:
            contact.assigned_counselor_id = payload.assigned_counselor_id
        if old_counselor != contact.assigned_counselor_id:
            assigned_cns = db.query(Counselor).filter(Counselor.id == contact.assigned_counselor_id).first() if contact.assigned_counselor_id else None
            activity = CounselorActivity(
                contact_id=contact.id,
                counselor_id=counselor_id,
                action_type="Assignment",
                outcome=f"Assigned to {assigned_cns.name if assigned_cns else 'Unassigned'}",
                notes=f"Assigned counselor changed to {assigned_cns.name if assigned_cns else 'Unassigned'}"
            )
            db.add(activity)

    contact.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(contact)

    try:
        from src.events import event_manager
        event_manager.broadcast_sync(
            "CONTACT_UPDATED",
            {"contact_id": contact.id, "status": contact.status, "counselor_followup_status": contact.counselor_followup_status},
            school_id=contact.school_id
        )
    except Exception:
        pass
    return {
        "success": True, 
        "contact": {
            "id": contact.id, 
            "notes": contact.notes, 
            "email": contact.email, 
            "name": contact.name, 
            "status": contact.status,
            "assigned_counselor_id": contact.assigned_counselor_id,
            "counselor_followup_status": contact.counselor_followup_status
        }
    }
